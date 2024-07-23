import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def write_to_excel(file1_data, file2_data, title_row1, title_row2, differences, output_file_path, sheet1_name, sheet2_name, key_columns, file1_column_mapping, file2_column_mapping):
    # Create a new workbook and add sheets
    workbook = openpyxl.Workbook()
    
    # Summary Sheet
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    
    # Define styles
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Light gray for headers
    header_font = Font(bold=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    diff_fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")  # Light red for differences
    extra_fill = PatternFill(start_color="DDFFDD", end_color="DDFFDD", fill_type="solid")  # Light green for extra rows
    
    # Write header for summary sheet
    summary_sheet.append([])
    summary_sheet.append(["Differences"])
    summary_sheet.append(["Combined ID", "Column", "File 1 Value", "File 2 Value"])
    
    # Apply header formatting
    for cell in summary_sheet[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Write differences to the summary sheet
    for combined_id, column, value1, value2 in differences:
        combined_id_str = " | ".join(combined_id)
        summary_sheet.append([combined_id_str, column, value1, value2])
    
    # Add extra rows to the summary sheet
    file1_extra_rows = set(tuple(row[file1_column_mapping[col]] for col in key_columns) for row in file1_data)
    file2_extra_rows = set(tuple(row[file2_column_mapping[col]] for col in key_columns) for row in file2_data)
    
    extra_file1_rows = file1_extra_rows - file2_extra_rows
    extra_file2_rows = file2_extra_rows - file1_extra_rows
    
    if extra_file1_rows:
        summary_sheet.append([])
        summary_sheet.append([f"Extra Rows in {sheet1_name}"])
        summary_sheet.append(["Combined ID"] + [title_row1[col] for col in file1_column_mapping.values()])
        for row in extra_file1_rows:
            summary_sheet.append([f" | ".join(row)] + [''] * (len(title_row1) - 1))
    
    if extra_file2_rows:
        summary_sheet.append([])
        summary_sheet.append([f"Extra Rows in {sheet2_name}"])
        summary_sheet.append(["Combined ID"] + [title_row2[col] for col in file2_column_mapping.values()])
        for row in extra_file2_rows:
            summary_sheet.append([f" | ".join(row)] + [''] * (len(title_row2) - 1))
    
    # Add file1 data
    file1_sheet = workbook.create_sheet(title=sheet1_name)
    # Write headers
    file1_sheet.append(title_row1)
    for cell in file1_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Write data
    for row in file1_data:
        file1_sheet.append(row)
    
    # Add file2 data
    file2_sheet = workbook.create_sheet(title=sheet2_name)
    # Write headers
    file2_sheet.append(title_row2)
    for cell in file2_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # Write data
    for row in file2_data:
        file2_sheet.append(row)
    
    # Highlight differences in file1 and file2 sheets
    def highlight_differences(sheet, differences, column_mapping):
        for combined_id, column, value1, value2 in differences:
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column), start=2):
                row_id = tuple(row[column_mapping[col]].value for col in key_columns)
                if row_id == combined_id:
                    cell_idx = column_mapping[column]
                    row[cell_idx].fill = diff_fill
    
    # Highlight differences in both sheets
    highlight_differences(file1_sheet, differences, file1_column_mapping)
    highlight_differences(file2_sheet, differences, file2_column_mapping)
    
    # Highlight extra rows
    def highlight_extra_rows(sheet, extra_rows, column_mapping):
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
            row_id = tuple(row[column_mapping[col]].value for col in key_columns)
            if row_id in extra_rows:
                for cell in row:
                    cell.fill = extra_fill
    
    highlight_extra_rows(file1_sheet, extra_file1_rows, file1_column_mapping)
    highlight_extra_rows(file2_sheet, extra_file2_rows, file2_column_mapping)
    
    # Save the workbook
    workbook.save(output_file_path)
