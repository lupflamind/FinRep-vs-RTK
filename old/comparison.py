import csv
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def read_csv(file_path):
    with open(file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile, delimiter=';')
        rows = list(reader)
    return rows

def find_title_row(rows):
    for i, row in enumerate(rows):
        if "Deal Id" in row:
            return i, row
    return None, None

def clean_rows(rows, title_row):
    clean_data = []
    title_length = len(title_row)
    
    for row in rows:
        if len(row) == title_length and any(row):
            clean_data.append(row)
    return clean_data

def remove_totals(rows, title_row):
    deal_id_index = title_row.index("Deal Id")
    filtered_data = [row for row in rows if row[deal_id_index].strip()]
    return filtered_data

def create_header_index_map(title_row):
    return {header: index for index, header in enumerate(title_row)}

def compare_files(file1_data, file2_data, file1_headers, file2_headers):
    differences = []

    file1_header_map = create_header_index_map(file1_headers)
    file2_header_map = create_header_index_map(file2_headers)

    common_headers = set(file1_header_map.keys()) & set(file2_header_map.keys())
    
    file1_dict = {row[file1_header_map["Deal Id"]]: row for row in file1_data}
    file2_dict = {row[file2_header_map["Deal Id"]]: row for row in file2_data}
    
    for deal_id in file1_dict:
        if deal_id in file2_dict:
            row1 = file1_dict[deal_id]
            row2 = file2_dict[deal_id]
            for header in common_headers:
                index1 = file1_header_map[header]
                index2 = file2_header_map[header]
                if row1[index1] != row2[index2]:
                    differences.append((deal_id, header, row1[index1], row2[index2]))
                    
    return differences

def write_csv(file_path, title_row, data):
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile, delimiter=';')
        writer.writerow(title_row)
        writer.writerows(data)

def write_to_excel(file1_filtered, file2_filtered, title_row1, title_row2, differences, prefix, suffix1, suffix2):
    wb = Workbook()

    # Write File 1 data
    ws1 = wb.active
    ws1.title = suffix1
    ws1.append(title_row1)
    for row in file1_filtered:
        ws1.append(row)

    # Write File 2 data
    ws2 = wb.create_sheet(title=suffix2)
    ws2.append(title_row2)
    for row in file2_filtered:
        ws2.append(row)

    # Highlight differences
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    file1_header_map = create_header_index_map(title_row1)
    file2_header_map = create_header_index_map(title_row2)
    
    file1_deal_ids = [row[file1_header_map["Deal Id"]] for row in file1_filtered]
    file2_deal_ids = [row[file2_header_map["Deal Id"]] for row in file2_filtered]

    for diff in differences:
        deal_id, column, file1_value, file2_value = diff
        if deal_id in file1_deal_ids:
            row_index1 = file1_deal_ids.index(deal_id) + 2  # +2 to account for header and 1-based index
            col_index1 = file1_header_map[column] + 1  # +1 to convert 0-based index to 1-based index
            ws1.cell(row=row_index1, column=col_index1).fill = highlight_fill

        if deal_id in file2_deal_ids:
            row_index2 = file2_deal_ids.index(deal_id) + 2  # +2 to account for header and 1-based index
            col_index2 = file2_header_map[column] + 1
            ws2.cell(row=row_index2, column=col_index2).fill = highlight_fill

    excel_filename = f"{prefix}.xlsx"
    wb.save(excel_filename)

# File paths
file1_path = 'fin-rep-vs-RTK/SpPos_et_RTK.txt'
file2_path = 'fin-rep-vs-RTK/SpPos_et_FinRep.txt'

# Extract prefix and suffix
prefix = os.path.basename(file1_path).split('_')[0] + '_' + os.path.basename(file1_path).split('_')[1]
suffix1 = os.path.basename(file1_path).split('_')[-1].split('.')[0]
suffix2 = os.path.basename(file2_path).split('_')[-1].split('.')[0]

# Read files
file1_rows = read_csv(file1_path)
file2_rows = read_csv(file2_path)

# Identify title rows
title_index1, title_row1 = find_title_row(file1_rows)
title_index2, title_row2 = find_title_row(file2_rows)

# Clean rows
file1_cleaned = clean_rows(file1_rows[title_index1+1:], title_row1)
file2_cleaned = clean_rows(file2_rows[title_index2+1:], title_row2)

# Remove totals
file1_filtered = remove_totals(file1_cleaned, title_row1)
file2_filtered = remove_totals(file2_cleaned, title_row2)

# Write filtered data to new CSV files
write_csv('file1_filtered.csv', title_row1, file1_filtered)
write_csv('file2_filtered.csv', title_row2, file2_filtered)

# Compare files
differences = compare_files(file1_filtered, file2_filtered, title_row1, title_row2)

# Output differences to Excel
write_to_excel(file1_filtered, file2_filtered, title_row1, title_row2, differences, prefix, suffix1, suffix2)

# Output differences to console
for diff in differences:
    print(f"Deal ID: {diff[0]}, Column: {diff[1]}, File1: {diff[2]}, File2: {diff[3]}")
